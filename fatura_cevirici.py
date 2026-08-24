# -*- coding: utf-8 -*-
"""
XML FATURA AKTARICI  —  Profesyonel Sürüm
==========================================================
GİB e-Fatura / e-Arşiv (UBL-TR 1.2) XML dosyalarını Excel listesine çevirir.

Çıktı Modları
-------------
Tekli Mod     : Her fatura = 1 satır  (muhasebe sisteminin kabul ettiği format)
Ayrıntılı Mod : Her kalem  = 1 satır  (kalem bazlı analiz için)

Tekli Mod çıktısı, sistemin kabul ettiği örnek listeyle birebir aynıdır:
  Fatura Tarihi | Fatura Numarası | Firma Ünvanı | Vergi Kimlik Numarası |
  Malın Cinsi | Miktar | Fatura Matrahı | Fatura KDV Tutarı | Tevkifat Tutarı

Bu sürümde düzeltilen kritik hatalar
------------------------------------
  1. KDV hesabı  : Önceki sürüm ÖİV / Telsiz / Damga gibi KDV DIŞI vergileri
                   de "Fatura KDV Tutarı"na ekliyordu (TaxScheme yanlış yerde
                   aranıyordu). Artık yalnızca gerçek KDV (kod 0015) alınır.
  2. Vergi No    : Önceki sürüm ilk "PartyIdentification"ı alıyordu; bu bazen
                   Ticaret Sicil / MERSİS No oluyordu. Artık schemeID=VKN/TCKN
                   olan kimlik, 10/11 haneli doğrulamayla seçilir.
  3. Firma Adı   : Artık "PartyLegalEntity/RegistrationName" da kontrol edilir
                   (kurumsal faturaların çoğunda ünvan buradadır).
  4. Miktar/Birim: "1" ve "ADET" ayrı düşmüyor; tek hücrede "1 ADET" olarak,
                   birden çok birim varsa "9 ADET, 3 METRE" biçiminde birleşir.
  5. Birim kodu  : Bilinmeyen/standart dışı kodlar (T0 vb.) akıllıca çözülür.
  6. Boşluk/satır: Ünvan ve açıklamalardaki gizli satır sonları temizlenir.

AI Özellikleri (opsiyonel — VARSAYILAN KAPALI, Anthropic API Key gerekir)
  • Cins Gruplama      • Firma Normalizasyonu
  • Birim Normalizasyonu • Bozuk XML Kurtarma • Anomali Tespiti
"""

import os
import re
import json
import threading
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from collections import OrderedDict

import pandas as pd

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk, scrolledtext
    _GUI_AVAILABLE = True
except Exception:                      # tkinter olmayan ortamlarda (test vb.)
    _GUI_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════
#  1) METİN & SAYI YARDIMCILARI
# ═══════════════════════════════════════════════════════════════════

_WS_RE = re.compile(r"\s+")


def clean_text(value) -> str:
    """Baştaki/sondaki boşlukları atar, iç satır sonları/çoklu boşlukları
    tek boşluğa indirger. 'Adet' ve '1'in ayrı satıra düşmesini engeller."""
    if value is None:
        return ""
    return _WS_RE.sub(" ", str(value)).strip()


def fmt_qty(val) -> str:
    """Miktarı okunaklı yazar: tamsa '13', ondalıksa nokta ile '1987.03'."""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return clean_text(val)
    if f == int(f):
        return str(int(f))
    return f"{f:.4f}".rstrip("0").rstrip(".")


def try_decode(data):
    if data is None:
        return None
    for enc in ("utf-8", "utf-8-sig", "iso-8859-9", "cp1254", "utf-16", "latin-1"):
        try:
            return data.decode(enc)
        except (UnicodeDecodeError, AttributeError):
            continue
    return None


def read_file(path):
    try:
        with open(path, "rb") as f:
            raw = f.read()
        if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
            return raw.decode("utf-16")
        if raw.startswith(b"\xef\xbb\xbf"):
            return raw.decode("utf-8-sig")
        return try_decode(raw)
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════
#  2) NAMESPACE-BAĞIMSIZ XML YARDIMCILARI
#     (cac/cbc önek adına bakmadan, yerel etiket adıyla çalışır)
# ═══════════════════════════════════════════════════════════════════

def _lname(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _kids(el, name):
    """el'in yalnızca DOĞRUDAN çocukları arasında yerel adı `name` olanlar."""
    if el is None:
        return []
    return [c for c in el if _lname(c.tag) == name]


def _first(el, name):
    for c in _kids(el, name):
        return c
    return None


def _path(el, *names):
    """Ardışık doğrudan-çocuk adımlarını izler: _path(root, 'Party', 'PartyName')."""
    cur = [el] if el is not None else []
    for name in names:
        nxt = []
        for e in cur:
            nxt.extend(_kids(e, name))
        cur = nxt
    return cur


def _desc(el, name):
    """el'in TÜM alt ağacında yerel adı `name` olan öğeler."""
    if el is None:
        return []
    return [e for e in el.iter() if _lname(e.tag) == name]


def _text_of(el):
    return clean_text(el.text) if (el is not None and el.text) else ""


def _first_text(el, *names):
    """İç içe doğrudan-çocuk yolundan ilk dolu metni döndürür."""
    for node in _path(el, *names):
        t = _text_of(node)
        if t:
            return t
    return ""


# ═══════════════════════════════════════════════════════════════════
#  3) BİRİM KODU → TÜRKÇE BİRİM ADI
# ═══════════════════════════════════════════════════════════════════

# GİB / UN-ECE Rec.20 birim kodları → sade Türkçe ad
UNIT_MAP = {
    # ADET
    "C62": "ADET", "H87": "ADET", "NIU": "ADET", "PCE": "ADET", "EA": "ADET",
    "NMP": "ADET", "NAR": "ADET", "PR": "ÇİFT", "ZZ": "ADET", "XPP": "ADET",
    "T0": "ADET",            # Turkcell/telekom hizmet kalemleri
    # AĞIRLIK
    "KGM": "KG", "KG": "KG", "GRM": "GRAM", "MGM": "MG",
    "TNE": "TON", "DTN": "TON",
    # UZUNLUK / ALAN / HACİM
    "MTR": "METRE", "MTK": "M2", "MTQ": "M3", "CMT": "CM", "KTM": "KM",
    "MMT": "MM",
    # HACİM (sıvı)
    "LTR": "LİTRE", "MLT": "ML",
    # ZAMAN
    "HUR": "SAAT", "MIN": "DAKİKA", "D33": "GÜN", "DAY": "GÜN", "DAE": "GÜN",
    "MON": "AY", "ANN": "YIL", "WEE": "HAFTA",
    # ENERJİ
    "KWH": "KWH", "MWH": "MWH", "GWH": "GWH",
    # PAKETLEME / DİĞER
    "BX": "KUTU", "CT": "KOLİ", "PA": "PAKET", "SET": "SET", "KT": "SET",
    "DZN": "DÜZINE", "ROL": "RULO", "BG": "TORBA",
}

# Türkçe yazımların normalizasyonu (aynı birimin farklı yazımlarını birleştirir)
BIRIM_NORM = {
    "AD": "ADET", "ADE": "ADET", "ADET": "ADET", "AGET": "ADET",
    "PCS": "ADET", "PC": "ADET", "PIECE": "ADET",
    "KG": "KG", "KGS": "KG", "KILOGRAM": "KG",
    "GR": "GRAM", "GRM": "GRAM", "GRAM": "GRAM", "MG": "MG",
    "TON": "TON", "TN": "TON",
    "LT": "LİTRE", "LTR": "LİTRE", "LİTRE": "LİTRE", "LITRE": "LİTRE",
    "LITER": "LİTRE", "ML": "ML",
    "MT": "METRE", "M": "METRE", "METRE": "METRE", "METER": "METRE",
    "M2": "M2", "M²": "M2", "MTK": "M2",
    "M3": "M3", "M³": "M3", "MTQ": "M3",
    "CM": "CM", "MM": "MM", "KM": "KM",
    "SAAT": "SAAT", "SA": "SAAT", "HR": "SAAT", "DAKİKA": "DAKİKA",
    "GÜN": "GÜN", "GUN": "GÜN", "DAY": "GÜN",
    "AY": "AY", "MONTH": "AY", "YIL": "YIL", "YL": "YIL", "YEAR": "YIL",
    "KUTU": "KUTU", "KT": "SET", "BOX": "KUTU", "KOLİ": "KOLİ", "KOLI": "KOLİ",
    "PAKET": "PAKET", "SET": "SET", "RULO": "RULO", "TORBA": "TORBA",
    "ÇİFT": "ÇİFT", "CIFT": "ÇİFT", "PR": "ÇİFT",
    "DÜZINE": "DÜZINE", "DUZINE": "DÜZINE",
    "KWH": "KWH", "MWH": "MWH", "GWH": "GWH",
}

BILINEN_BIRIM_ADLARI = set(UNIT_MAP.values()) | set(BIRIM_NORM.values())

# Bir birim kodu bilinmediğinde varsayılan (hizmet/genel kalemler "1 ADET").
VARSAYILAN_BIRIM = "ADET"


def normalize_birim(birim_adi: str) -> str:
    if not birim_adi:
        return ""
    key = birim_adi.strip().upper()
    return BIRIM_NORM.get(key, key)


def get_unit_name(code: str) -> str:
    """XML unitCode → Türkçe birim adı. Bilinmiyorsa VARSAYILAN_BIRIM."""
    if not code:
        return VARSAYILAN_BIRIM
    key = code.strip().upper()
    if key in UNIT_MAP:
        return UNIT_MAP[key]
    if key in BIRIM_NORM:
        return BIRIM_NORM[key]
    # Standart dışı / bilinmeyen kod → adet varsay (hizmet kalemleri)
    return VARSAYILAN_BIRIM


# ═══════════════════════════════════════════════════════════════════
#  4) DÖVİZ
# ═══════════════════════════════════════════════════════════════════

def get_document_currency(root):
    t = _first_text(root, "DocumentCurrencyCode")
    return t.upper() if t else "TRY"


def get_exchange_rate(root):
    for tag in ("TaxExchangeRate", "PricingExchangeRate"):
        for rate_elem in _kids(root, tag):
            calc = _first(rate_elem, "CalculationRate")
            if calc is not None and calc.text:
                try:
                    r = float(calc.text)
                    if r > 0:
                        return r
                except (ValueError, TypeError):
                    pass
    return 1.0


def _amount_to_try(elem, invoice_currency, exchange_rate):
    """Tek bir tutar öğesini TRY'ye çevirir."""
    if elem is None or not elem.text:
        return 0.0
    try:
        val = float(elem.text)
    except (TypeError, ValueError):
        return 0.0
    curr = (elem.get("currencyID") or invoice_currency).strip().upper()
    return val if curr == "TRY" else val * exchange_rate


def _sum_amounts_try(elems, invoice_currency, exchange_rate):
    """Birden çok tutar öğesinden ilk uygun olanı TRY döndürür (öncelik TRY)."""
    elems = list(elems)
    for e in elems:
        if (e.get("currencyID") or "").strip().upper() == "TRY" and e.text:
            try:
                return float(e.text)
            except (ValueError, TypeError):
                pass
    for e in elems:
        if e.text:
            return _amount_to_try(e, invoice_currency, exchange_rate)
    return 0.0


# ═══════════════════════════════════════════════════════════════════
#  5) FATURA TİPİ
# ═══════════════════════════════════════════════════════════════════

FATURA_TIP_MAP = {
    "SATIS": "Satış", "IADE": "İade", "TEVKIFAT": "Tevkifat",
    "ISTISNA": "İstisna", "OZELMATRAH": "Özel Matrah", "IHRACAT": "İhracat",
    "IHRACKAYITLI": "İhraç Kayıtlı", "MUSTAHSIL": "Müstahsil",
    "KOMISYONCU": "Komisyoncu", "IPTAL": "İptal", "SGK": "SGK",
}


def get_invoice_type(root):
    kod = _first_text(root, "InvoiceTypeCode").upper() or "SATIS"
    return kod, FATURA_TIP_MAP.get(kod, kod.title())


# ═══════════════════════════════════════════════════════════════════
#  6) TARAF (SATICI) BİLGİLERİ — VKN & ÜNVAN
# ═══════════════════════════════════════════════════════════════════

def _extract_vkn(party):
    """Satıcının Vergi Kimlik / TC Kimlik numarasını güvenilir seçer.
    schemeID=VKN/TCKN önceliklidir; yoksa 10/11 haneli sayısal kimlik."""
    if party is None:
        return ""
    ids = []                      # (schemeID_upper, value)
    for pid in _kids(party, "PartyIdentification"):
        id_el = _first(pid, "ID")
        if id_el is None:
            continue
        val = clean_text(id_el.text)
        scheme = (id_el.get("schemeID") or "").strip().upper()
        if val:
            ids.append((scheme, val))

    # 1) Doğrudan VKN / TCKN şeması
    for want in ("VKN", "TCKN"):
        for scheme, val in ids:
            if scheme == want:
                return val
    # 2) TC Kimlik'in diğer yazımı
    for scheme, val in ids:
        if scheme in ("TC", "TCKIMLIK", "TCKIMLIKNO"):
            return val
    # 3) Şema yok ama 10 (VKN) / 11 (TCKN) haneli sayı
    for scheme, val in ids:
        digits = re.sub(r"\D", "", val)
        if len(digits) in (10, 11) and digits == val:
            return val
    # 4) PartyTaxScheme altındaki kimlik (bazı faturalarda VKN oradadır)
    for tax in _kids(party, "PartyTaxScheme"):
        cid = _first_text(tax, "CompanyID")
        d = re.sub(r"\D", "", cid)
        if len(d) in (10, 11):
            return cid
    # 5) Son çare: ilk geçerli görünen kimlik
    for scheme, val in ids:
        if scheme not in ("BAYINO", "HIZMETNO"):
            return val
    return ids[0][1] if ids else ""


def _extract_unvan(party):
    """Satıcı ünvanını en resmi kaynaktan alır.
    Öncelik: PartyLegalEntity/RegistrationName → PartyName/Name → Person."""
    if party is None:
        return ""
    reg = _first_text(party, "PartyLegalEntity", "RegistrationName")
    if reg:
        return reg
    name = _first_text(party, "PartyName", "Name")
    if name:
        return name
    fn = _first_text(party, "Person", "FirstName")
    mn = _first_text(party, "Person", "MiddleName")
    ln = _first_text(party, "Person", "FamilyName")
    person = clean_text(" ".join(p for p in (fn, mn, ln) if p))
    return person


# ═══════════════════════════════════════════════════════════════════
#  7) VERGİLER — KDV & TEVKİFAT
# ═══════════════════════════════════════════════════════════════════

KDV_KODU = "0015"


def _tax_type_code(subtotal):
    """TaxSubtotal içindeki vergi tür kodunu (0015 = KDV) bulur.
    TaxTypeCode, TaxCategory/TaxScheme altında iç içe olabilir — alt ağaçta arar."""
    for tc in _desc(subtotal, "TaxTypeCode"):
        if tc.text and tc.text.strip():
            return tc.text.strip()
    # Kod yoksa vergi adından çıkarım
    for nm in _desc(subtotal, "Name"):
        up = (nm.text or "").upper()
        if "KATMA DE" in up or up.strip() in ("KDV", "K.D.V."):
            return KDV_KODU
    return ""


def _collect_kdv(scope, inv_cur, rate):
    """scope (fatura kökü veya kalem) altındaki TaxTotal'lardan yalnızca
    gerçek KDV'yi (0015) toplar. ÖİV/ÖTV/Telsiz/Damga hariç tutulur."""
    kdv = 0.0
    subtotal_count = 0
    for tt in _kids(scope, "TaxTotal"):
        for sub in _desc(tt, "TaxSubtotal"):
            subtotal_count += 1
            amt = _sum_amounts_try(_kids(sub, "TaxAmount"), inv_cur, rate)
            code = _tax_type_code(sub)
            if code == KDV_KODU:
                kdv += amt
            elif code == "":
                # Kod hiç yok: tek satırlıysa KDV varsay (eski faturalar)
                if subtotal_count == 1 and len(_desc(tt, "TaxSubtotal")) == 1:
                    kdv += amt
    return kdv


def _collect_tevkifat(scope, inv_cur, rate):
    """WithholdingTaxTotal (KDV tevkifatı) toplamını döndürür."""
    tev = 0.0
    for wtt in _kids(scope, "WithholdingTaxTotal"):
        direct = _kids(wtt, "TaxAmount")     # rollup toplamı
        if direct:
            tev += _sum_amounts_try(direct, inv_cur, rate)
        else:
            for sub in _desc(wtt, "TaxSubtotal"):
                tev += _sum_amounts_try(_kids(sub, "TaxAmount"), inv_cur, rate)
    return tev


# ═══════════════════════════════════════════════════════════════════
#  8) XML PARSE
# ═══════════════════════════════════════════════════════════════════

def _get_currency_info(root):
    inv_cur = get_document_currency(root)
    rate = get_exchange_rate(root) if inv_cur != "TRY" else 1.0
    return inv_cur, rate


def parse_header(root, inv_cur="TRY", rate=1.0):
    tip_kodu, tip_adi = get_invoice_type(root)

    fatura_no = _first_text(root, "ID")
    fatura_tarihi = _first_text(root, "IssueDate")

    supplier = _first(root, "AccountingSupplierParty")
    party = _first(supplier, "Party") if supplier is not None else None
    firma_unvani = _extract_unvan(party)
    vkn_tckn = _extract_vkn(party)

    # Matrah = KDV hariç toplam (LegalMonetaryTotal/TaxExclusiveAmount)
    toplam_matrah = 0.0
    legal = _first(root, "LegalMonetaryTotal")
    if legal is not None:
        te = _first(legal, "TaxExclusiveAmount")
        toplam_matrah = _amount_to_try(te, inv_cur, rate)
    if toplam_matrah == 0.0:
        for line in _desc(root, "InvoiceLine"):
            toplam_matrah += _amount_to_try(_first(line, "LineExtensionAmount"), inv_cur, rate)

    # KDV & Tevkifat (fatura başı)
    toplam_kdv = _collect_kdv(root, inv_cur, rate)
    toplam_tevkifat = _collect_tevkifat(root, inv_cur, rate)

    # Fatura başında yoksa kalem düzeyinden topla
    if toplam_kdv == 0.0:
        for line in _desc(root, "InvoiceLine"):
            toplam_kdv += _collect_kdv(line, inv_cur, rate)
    if toplam_tevkifat == 0.0:
        for line in _desc(root, "InvoiceLine"):
            toplam_tevkifat += _collect_tevkifat(line, inv_cur, rate)

    return {
        "fatura_no": fatura_no,
        "fatura_tarihi": fatura_tarihi,
        "fatura_tip_kodu": tip_kodu,
        "fatura_tip_adi": tip_adi,
        "firma_unvani": firma_unvani,
        "vkn_tckn": vkn_tckn,
        "toplam_matrah": round(toplam_matrah, 2),
        "toplam_kdv": round(toplam_kdv, 2),
        "toplam_tevkifat": round(toplam_tevkifat, 2),
        "para_birimi": inv_cur,
        "kur": round(rate, 6) if inv_cur != "TRY" else None,
    }


def parse_lines(root, inv_cur="TRY", rate=1.0):
    result = []
    for line in _desc(root, "InvoiceLine"):
        item = _first(line, "Item")
        mal_cinsi = _first_text(item, "Name") if item is not None else ""

        miktar = 0.0
        birim_kodu = ""
        qty = _first(line, "InvoicedQuantity")
        if qty is not None:
            try:
                miktar = float(qty.text)
            except (TypeError, ValueError):
                miktar = 0.0
            birim_kodu = (qty.get("unitCode") or "").strip()

        birim_fiyat = _amount_to_try(
            _first(_first(line, "Price"), "PriceAmount") if _first(line, "Price") is not None else None,
            inv_cur, rate)

        matrah = _amount_to_try(_first(line, "LineExtensionAmount"), inv_cur, rate)
        kdv = _collect_kdv(line, inv_cur, rate)
        tevkifat = _collect_tevkifat(line, inv_cur, rate)

        result.append({
            "mal_cinsi": clean_text(mal_cinsi),
            "miktar": miktar,
            "birim_kodu": birim_kodu,
            "birim": get_unit_name(birim_kodu),
            "birim_fiyat": round(birim_fiyat, 4),
            "matrah": round(matrah, 2),
            "kdv": round(kdv, 2),
            "tevkifat": round(tevkifat, 2),
        })
    return result


# ── Malın Cinsi & Miktar oluşturucular (AI yeniden yazımı da kullanır) ──

def build_malin_cinsi(kalemler, cins_map=None):
    """Kalem adlarını tekilleştirip ' - ' ile birleştirir.
    cins_map verilirse (AI cins gruplama) adlar önce eşlenir."""
    names = []
    for k in kalemler:
        ham = k["mal_cinsi"] or "(İsimsiz)"
        ad = clean_text((cins_map or {}).get(ham, ham)) or "(İsimsiz)"
        if ad not in names:
            names.append(ad)
    return " - ".join(names) if names else "(İsimsiz)"


def build_miktar(kalemler, birim_override=None):
    """Birim bazında miktarları toplayıp 'qty BİRİM' biçiminde birleştirir.
    Örn: {ADET:9, METRE:3} → '9 ADET, 3 METRE'.
    birim_override: {ham_kod: birim_adi} (AI birim normalizasyonu için)."""
    groups = OrderedDict()
    for k in kalemler:
        birim = ""
        if birim_override and k["birim_kodu"] in birim_override:
            birim = normalize_birim(birim_override[k["birim_kodu"]])
        if not birim:
            birim = normalize_birim(k["birim"]) or VARSAYILAN_BIRIM
        groups[birim] = groups.get(birim, 0.0) + (k["miktar"] or 0.0)
    return ", ".join(f"{fmt_qty(q)} {b}" for b, q in groups.items())


def _doviz_notu(h):
    ic = h["para_birimi"]
    if ic == "TRY":
        return ""
    if h["kur"]:
        return f" [{ic} @ {h['kur']} → TRY]"
    return f" [{ic} — kur yok]"


# ═══════════════════════════════════════════════════════════════════
#  9) MOD 1 — TEKLİ  (sistemin kabul ettiği 9 sütunluk format)
# ═══════════════════════════════════════════════════════════════════

def parse_tekli(xml_content, filename, gruplama_modu="tam", api_key=""):
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        print(f"  ⚠ XML Parse Hatası ({filename}): {e}")
        return []
    except Exception as e:
        print(f"  ⚠ Parse Hatası ({filename}): {e}")
        return []

    try:
        inv_cur, rate = _get_currency_info(root)
        h = parse_header(root, inv_cur, rate)
        kalemler = parse_lines(root, inv_cur, rate)
        if not kalemler:
            return []

        # "ilk_kelime" modu: adları ilk kelimeye indirge (hızlı, AI'sız)
        cins_map = None
        if gruplama_modu == "ilk_kelime":
            cins_map = {}
            for k in kalemler:
                ham = k["mal_cinsi"] or "(İsimsiz)"
                cins_map[ham] = (ham.split() or ["(İsimsiz)"])[0]

        return [{
            "Fatura Tarihi":         h["fatura_tarihi"],
            "Fatura Numarası":       h["fatura_no"],
            "Firma Ünvanı":          h["firma_unvani"],
            "Vergi Kimlik Numarası": h["vkn_tckn"],
            "Malın Cinsi":           build_malin_cinsi(kalemler, cins_map),
            "Miktar":                build_miktar(kalemler),
            "Fatura Matrahı":        h["toplam_matrah"],
            "Fatura KDV Tutarı":     h["toplam_kdv"],
            "Tevkifat Tutarı":       h["toplam_tevkifat"],
            # Yardımcı/gizli alanlar (Excel'e yazılmadan önce temizlenir)
            "_sheet":                "Satış Faturaları",
            "_doviz_notu":           _doviz_notu(h) + (
                                        f" [{h['fatura_tip_adi']}]"
                                        if h["fatura_tip_kodu"] != "SATIS" else ""),
            "_kalemler":             kalemler,   # AI cins/birim yeniden yazımı için
        }]
    except Exception as e:
        print(f"  ⚠ Parse Hatası ({filename}): {e}")
        return []


# ═══════════════════════════════════════════════════════════════════
#  10) MOD 2 — AYRINTILI  (her kalem = 1 satır)
# ═══════════════════════════════════════════════════════════════════

def parse_ayrintili(xml_content, filename, **kwargs):
    try:
        root = ET.fromstring(xml_content)
    except Exception as e:
        print(f"  ⚠ Parse Hatası ({filename}): {e}")
        return []
    try:
        inv_cur, rate = _get_currency_info(root)
        h = parse_header(root, inv_cur, rate)
        kalemler = parse_lines(root, inv_cur, rate)
        if not kalemler:
            return []

        rows = []
        for k in kalemler:
            birim = normalize_birim(k["birim"]) or VARSAYILAN_BIRIM
            rows.append({
                "Fatura Tarihi":          h["fatura_tarihi"],
                "Fatura Numarası":        h["fatura_no"],
                "Firma Ünvanı":           h["firma_unvani"],
                "Vergi Kimlik Numarası":  h["vkn_tckn"],
                "Malın Cinsi":            k["mal_cinsi"] or "(İsimsiz)",
                "Miktar":                 f"{fmt_qty(k['miktar'])} {birim}",
                "Birim Fiyatı":           k["birim_fiyat"],
                "Kalem Matrahı":          k["matrah"],
                "Kalem KDV Tutarı":       k["kdv"],
                "Kalem Tevkifat":         k["tevkifat"],
                "Fatura Matrahı":         h["toplam_matrah"],
                "Fatura KDV Tutarı":      h["toplam_kdv"],
                "Fatura Tevkifat Tutarı": h["toplam_tevkifat"],
                "_sheet":                 "Fatura Detayı",
                "_doviz_notu":            _doviz_notu(h),
                "_ham_birim_kodu":        k["birim_kodu"],
                "_mal_cinsi_ham":         k["mal_cinsi"],
            })
        return rows
    except Exception as e:
        print(f"  ⚠ Parse Hatası ({filename}): {e}")
        return []


# ═══════════════════════════════════════════════════════════════════
#  11) AI ALTYAPI  (opsiyonel — varsayılan kapalı)
# ═══════════════════════════════════════════════════════════════════

_ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
_MODEL = "claude-sonnet-4-20250514"     # gerekirse güncel model kimliğiyle değiştirin
_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(.*?)```", re.DOTALL)


def _ai_json_call(prompt: str, api_key: str, max_tokens: int = 1024):
    payload = json.dumps({
        "model": _MODEL,
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        _ANTHROPIC_URL, data=payload,
        headers={
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
        })
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    raw = data["content"][0]["text"].strip()
    m = _JSON_FENCE_RE.search(raw)     # ```json ... ``` bloğunu düzgün ayıkla
    if m:
        raw = m.group(1).strip()
    return json.loads(raw)


def ai_grup_cinsi(mal_listesi, api_key):
    if not mal_listesi:
        return {}
    prompt = (
        "Fatura kalem adlarını ANA ÜRÜN CİNSİNE indirge.\n\n"
        "KURALLAR:\n"
        "1. Boyut, ebat, ölçü, renk, kod, seri no, tarih, miktar → ATILACAK\n"
        "2. Sonuç 1-3 kelime olmalı (örn: 'BETON BORU', 'DEMİR', 'HİZMET')\n"
        "3. Aynı ürün grubundaki tüm varyantlar → AYNI cins adı\n"
        "4. Türkçe büyük harf kullan\n\n"
        "YALNIZCA JSON döndür. Format: {\"orijinal\": \"CİNS\", ...}\n\n"
        "Kalemler:\n" + "\n".join(f"- {m}" for m in mal_listesi)
    )
    return _ai_json_call(prompt, api_key, max_tokens=2048) or {}


def ai_normalize_firmalar(firma_listesi, api_key):
    if not firma_listesi:
        return {}
    prompt = (
        "Aşağıdaki fatura firma ünvanları listesinde aynı firmaya ait farklı "
        "yazımları en tam ve doğru haliyle normalize et. Farklı firmalar farklı "
        "kalsın. YALNIZCA JSON döndür.\n"
        "Format: {\"orijinal\": \"normalize\", ...}\n\n"
        "Firmalar:\n" + "\n".join(f"- {f}" for f in firma_listesi)
    )
    return _ai_json_call(prompt, api_key, max_tokens=2048) or {}


def ai_normalize_birimler(birim_kontekst, api_key):
    if not birim_kontekst:
        return {}
    items = "\n".join(f"- Kod: {b['kod']}, Ürün: {b['mal_cinsi']}" for b in birim_kontekst)
    prompt = (
        "Fatura birim kodlarını Türkçe birim adına çevir. Ürün bağlamını kullan.\n"
        "Yaygın: ADET, KG, TON, LİTRE, METRE, M2, M3, KUTU, SET, SAAT, GÜN, KWH\n"
        "YALNIZCA JSON döndür. Format: {\"KOD\": \"BİRİM\", ...}\n\n"
        "Birimler:\n" + items
    )
    return _ai_json_call(prompt, api_key) or {}


def ai_recover_xml(xml_content, filename, api_key):
    prompt = (
        "Bozuk/standart dışı fatura XML'inden şu alanları çıkar. Bulamazsan null. "
        "YALNIZCA JSON döndür.\n"
        "{\"fatura_no\":null,\"fatura_tarihi\":null,\"firma_unvani\":null,"
        "\"vkn_tckn\":null,\"toplam_matrah\":null,\"toplam_kdv\":null,"
        "\"toplam_tevkifat\":null,\"mal_cinsi\":null}\n\nXML:\n" + xml_content[:4000]
    )
    return _ai_json_call(prompt, api_key) or {}


def ai_anomali_tespit(fatura_ozet, api_key):
    if not fatura_ozet:
        return []
    prompt = (
        "Aşağıdaki fatura listesini muhasebe açısından incele. Şüpheli/hatalı "
        "olanları tespit et. Kontrol: KDV oranı %0 veya >%25 mi? Matrah sıfır mı? "
        "Tevkifat KDV'den büyük mü? Tutar olağandışı mı?\n"
        "YALNIZCA anomali olanları döndür. Yoksa boş liste. YALNIZCA JSON.\n"
        "Format: [{\"fatura_no\":\"...\",\"anomali\":\"kısa açıklama\"}, ...]\n\n"
        "Faturalar:\n" + json.dumps(fatura_ozet, ensure_ascii=False)
    )
    r = _ai_json_call(prompt, api_key, max_tokens=2048)
    return r if isinstance(r, list) else []


# ═══════════════════════════════════════════════════════════════════
#  12) DOSYA TARAMA & İŞLEME
# ═══════════════════════════════════════════════════════════════════

def collect_xml_files(klasor_yolu):
    jobs = []
    for root_dir, dirs, files in os.walk(klasor_yolu, followlinks=True):
        dirs[:] = [d for d in dirs if not d.startswith(".") and d != "__MACOSX"]
        rel = os.path.relpath(root_dir, klasor_yolu)
        for file in files:
            path = os.path.join(root_dir, file)
            disp = f"{rel}/{file}" if rel != "." else file
            low = file.lower()
            if low.endswith(".xml"):
                jobs.append(("xml", path, disp))
            elif low.endswith(".zip"):
                jobs.append(("zip", path, disp))
    return jobs


def _ai_recover_row(content, display, api_key, mod):
    try:
        data = ai_recover_xml(content, display, api_key)
        if not data or not any(data.values()):
            return []
        base = {
            "Fatura Tarihi":         data.get("fatura_tarihi") or "",
            "Fatura Numarası":       data.get("fatura_no") or "",
            "Firma Ünvanı":          data.get("firma_unvani") or "",
            "Vergi Kimlik Numarası": data.get("vkn_tckn") or "",
            "Malın Cinsi":           data.get("mal_cinsi") or "(Kurtarıldı)",
            "Miktar":                "",
            "Fatura Matrahı":        data.get("toplam_matrah") or 0.0,
            "Fatura KDV Tutarı":     data.get("toplam_kdv") or 0.0,
            "Tevkifat Tutarı":       data.get("toplam_tevkifat") or 0.0,
            "_sheet":                "Satış Faturaları",
            "_doviz_notu":           "",
            "_kalemler":             [],
        }
        if mod == "ayrintili":
            base = {
                "Fatura Tarihi": base["Fatura Tarihi"],
                "Fatura Numarası": base["Fatura Numarası"],
                "Firma Ünvanı": base["Firma Ünvanı"],
                "Vergi Kimlik Numarası": base["Vergi Kimlik Numarası"],
                "Malın Cinsi": base["Malın Cinsi"],
                "Miktar": "",
                "Birim Fiyatı": 0.0, "Kalem Matrahı": 0.0,
                "Kalem KDV Tutarı": 0.0, "Kalem Tevkifat": 0.0,
                "Fatura Matrahı": base["Fatura Matrahı"],
                "Fatura KDV Tutarı": base["Fatura KDV Tutarı"],
                "Fatura Tevkifat Tutarı": base["Tevkifat Tutarı"],
                "_sheet": "Fatura Detayı", "_doviz_notu": "",
            }
        return [base]
    except Exception as e:
        print(f"  ⚠ AI kurtarma hatası ({display}): {e}")
        return []


def process_files(jobs, parse_fn, log_fn, progress_fn,
                  ai_xml_recovery=False, api_key="", mod="tekli"):
    all_data, basarili, basarisiz, kurtarilan = [], 0, 0, 0
    total = len(jobs)

    def _isle(content, display):
        nonlocal basarili, basarisiz, kurtarilan
        rows = parse_fn(content, display)
        if rows:
            note = rows[0].get("_doviz_notu", "")
            all_data.extend(rows)
            log_fn(f"✓ {display}{note}")
            basarili += 1
        elif ai_xml_recovery and api_key and "nvoice" in content:
            rec = _ai_recover_row(content, display, api_key, mod)
            if rec:
                all_data.extend(rec)
                log_fn(f"🤖 {display} — AI kurtarma başarılı")
                kurtarilan += 1
            else:
                log_fn(f"✗ {display} — kalem bulunamadı (AI kurtarma da başarısız)")
                basarisiz += 1
        else:
            log_fn(f"✗ {display} — kalem bulunamadı")
            basarisiz += 1

    for i, (tip, path, display) in enumerate(jobs):
        progress_fn(i + 1, total)
        if tip == "xml":
            content = read_file(path)
            if content:
                _isle(content, display)
            else:
                log_fn(f"✗ {display} — encoding hatası")
                basarisiz += 1
        elif tip == "zip":
            try:
                with zipfile.ZipFile(path, "r") as z:
                    xmls = [f for f in z.namelist()
                            if f.lower().endswith(".xml")
                            and not os.path.basename(f).startswith(".")]
                    if not xmls:
                        log_fn(f"— {display} (ZIP içinde XML yok)")
                        continue
                    for zf in xmls:
                        try:
                            zt = try_decode(z.read(zf))
                            if zt:
                                _isle(zt, f"{display} → {os.path.basename(zf)}")
                            else:
                                log_fn(f"✗ ZIP içi {zf} — encoding hatası")
                                basarisiz += 1
                        except Exception as e:
                            log_fn(f"✗ ZIP içi {zf} — {e}")
                            basarisiz += 1
            except zipfile.BadZipFile:
                log_fn(f"✗ {display} — geçersiz ZIP")
                basarisiz += 1
            except Exception as e:
                log_fn(f"✗ {display} — {e}")
                basarisiz += 1

    for row in all_data:
        row.pop("_doviz_notu", None)
    return all_data, basarili, basarisiz, kurtarilan


# ═══════════════════════════════════════════════════════════════════
#  13) AI SONRASI İŞLEME
# ═══════════════════════════════════════════════════════════════════

def post_process_ai(all_data, ai_features, api_key, mod, log_fn):
    _CINS_BATCH = 200
    _AI_BATCH = 60

    # ── Cins Gruplama (yalnızca Tekli) — toplu tek çağrı ──
    if ai_features.get("gruplama_modu") == "ai" and api_key and mod == "tekli":
        log_fn("🤖 [Cins] Kalem adları toplu analiz ediliyor...")
        tum = list(dict.fromkeys(
            k["mal_cinsi"] or "(İsimsiz)"
            for row in all_data for k in row.get("_kalemler", [])
        ))
        if tum:
            cins_map = {}
            try:
                for i in range(0, len(tum), _CINS_BATCH):
                    cins_map.update(ai_grup_cinsi(tum[i:i + _CINS_BATCH], api_key))
                for row in all_data:
                    if row.get("_kalemler"):
                        row["Malın Cinsi"] = build_malin_cinsi(row["_kalemler"], cins_map)
                log_fn(f"   ✓ {len(tum)} kalem → {len(set(cins_map.values()))} grup")
            except Exception as e:
                log_fn(f"   ⚠ Cins gruplama hatası: {e}")

    # ── Firma Normalizasyonu ──
    if ai_features.get("firma") and api_key:
        log_fn("🤖 [Firma] Ünvanlar normalize ediliyor...")
        firmalar = list(dict.fromkeys(
            r.get("Firma Ünvanı", "") for r in all_data if r.get("Firma Ünvanı")))
        try:
            fmap = ai_normalize_firmalar(firmalar, api_key)
            for row in all_data:
                o = row.get("Firma Ünvanı", "")
                row["Firma Ünvanı"] = fmap.get(o, o)
            log_fn(f"   ✓ {len(firmalar)} firma kontrol edildi")
        except Exception as e:
            log_fn(f"   ⚠ Firma normalizasyonu hatası: {e}")

    # ── Birim Normalizasyonu (bilinmeyen ham kodlar) ──
    if ai_features.get("birim") and api_key:
        log_fn("🤖 [Birim] Bilinmeyen birimler analiz ediliyor...")
        bilinmeyen = {}
        if mod == "tekli":
            for row in all_data:
                for k in row.get("_kalemler", []):
                    kod = k["birim_kodu"]
                    if kod and kod.strip().upper() not in UNIT_MAP \
                       and kod.strip().upper() not in BIRIM_NORM \
                       and kod not in bilinmeyen:
                        bilinmeyen[kod] = k["mal_cinsi"]
        else:
            for row in all_data:
                kod = row.get("_ham_birim_kodu", "")
                if kod and kod.strip().upper() not in UNIT_MAP \
                   and kod.strip().upper() not in BIRIM_NORM \
                   and kod not in bilinmeyen:
                    bilinmeyen[kod] = row.get("_mal_cinsi_ham", "")
        if bilinmeyen:
            try:
                bmap = ai_normalize_birimler(
                    [{"kod": k, "mal_cinsi": v} for k, v in bilinmeyen.items()], api_key)
                if mod == "tekli":
                    for row in all_data:
                        if row.get("_kalemler"):
                            row["Miktar"] = build_miktar(row["_kalemler"], bmap)
                else:
                    for row in all_data:
                        kod = row.get("_ham_birim_kodu", "")
                        if kod in bmap:
                            q = fmt_qty(row["Miktar"].split()[0]) if row.get("Miktar") else ""
                            row["Miktar"] = f"{q} {normalize_birim(bmap[kod])}".strip()
                log_fn(f"   ✓ {len(bilinmeyen)} bilinmeyen birim normalize edildi")
            except Exception as e:
                log_fn(f"   ⚠ Birim normalizasyonu hatası: {e}")
        else:
            log_fn("   ✓ Tüm birimler zaten tanımlı")

    # ── Anomali Tespiti ──
    anomali_toplam = 0
    if ai_features.get("anomali") and api_key:
        log_fn("🤖 [Anomali] Faturalar analiz ediliyor...")
        goruldu, ozet = set(), []
        for row in all_data:
            fn = row.get("Fatura Numarası", "")
            if fn and fn not in goruldu:
                goruldu.add(fn)
                ozet.append({
                    "fatura_no": fn,
                    "tarih": row.get("Fatura Tarihi", ""),
                    "firma": row.get("Firma Ünvanı", ""),
                    "matrah": float(row.get("Fatura Matrahı") or row.get("Kalem Matrahı") or 0),
                    "kdv": float(row.get("Fatura KDV Tutarı") or row.get("Kalem KDV Tutarı") or 0),
                    "tevkifat": float(row.get("Tevkifat Tutarı") or row.get("Fatura Tevkifat Tutarı") or 0),
                })
        anomali_map = {}
        for i in range(0, len(ozet), _AI_BATCH):
            try:
                for s in ai_anomali_tespit(ozet[i:i + _AI_BATCH], api_key):
                    fn, ac = s.get("fatura_no", ""), s.get("anomali", "")
                    if fn and ac:
                        anomali_map[fn] = ac
            except Exception as e:
                log_fn(f"   ⚠ Anomali batch {i // _AI_BATCH + 1} hatası: {e}")
        # Mükerrer fatura no (kural tabanlı)
        sayac = {}
        for row in all_data:
            fn = row.get("Fatura Numarası", "")
            if fn:
                sayac[fn] = sayac.get(fn, 0) + 1
        for fn, c in sayac.items():
            if c > 1:
                ek = f"Mükerrer fatura numarası ({c} kez)"
                anomali_map[fn] = (anomali_map[fn] + " | " + ek) if fn in anomali_map else ek
        for row in all_data:
            row["Anomali"] = anomali_map.get(row.get("Fatura Numarası", ""), "")
        anomali_toplam = len(anomali_map)
        log_fn(f"   ✓ {len(ozet)} fatura analiz edildi, {anomali_toplam} anomali")

    # Yardımcı alanları temizle
    for row in all_data:
        for k in ("_kalemler", "_ham_birim_kodu", "_mal_cinsi_ham"):
            row.pop(k, None)
    return all_data, anomali_toplam


# ═══════════════════════════════════════════════════════════════════
#  14) EXCEL KAYDET
# ═══════════════════════════════════════════════════════════════════

TEKLI_COLS = ["Fatura Tarihi", "Fatura Numarası", "Firma Ünvanı",
              "Vergi Kimlik Numarası", "Malın Cinsi", "Miktar",
              "Fatura Matrahı", "Fatura KDV Tutarı", "Tevkifat Tutarı"]

AYRINTILI_COLS = ["Fatura Tarihi", "Fatura Numarası", "Firma Ünvanı",
                  "Vergi Kimlik Numarası", "Malın Cinsi", "Miktar", "Birim Fiyatı",
                  "Kalem Matrahı", "Kalem KDV Tutarı", "Kalem Tevkifat",
                  "Fatura Matrahı", "Fatura KDV Tutarı", "Fatura Tevkifat Tutarı"]

TEKLI_WIDTHS = {"Fatura Tarihi": 14, "Fatura Numarası": 24, "Firma Ünvanı": 35,
                "Vergi Kimlik Numarası": 20, "Malın Cinsi": 52, "Miktar": 20,
                "Fatura Matrahı": 16, "Fatura KDV Tutarı": 18, "Tevkifat Tutarı": 16,
                "Anomali": 45}

AYRINTILI_WIDTHS = {"Fatura Tarihi": 14, "Fatura Numarası": 24, "Firma Ünvanı": 35,
                    "Vergi Kimlik Numarası": 20, "Malın Cinsi": 46, "Miktar": 14,
                    "Birim Fiyatı": 14, "Kalem Matrahı": 14, "Kalem KDV Tutarı": 15,
                    "Kalem Tevkifat": 14, "Fatura Matrahı": 16, "Fatura KDV Tutarı": 18,
                    "Fatura Tevkifat Tutarı": 18, "Anomali": 45}

MONEY_COLS = {"Fatura Matrahı", "Fatura KDV Tutarı", "Tevkifat Tutarı",
              "Fatura Tevkifat Tutarı", "Birim Fiyatı", "Kalem Matrahı",
              "Kalem KDV Tutarı", "Kalem Tevkifat"}

# Bu sütunlar metin olarak yazılır (baştaki sıfır korunur, tarih otomatik
# tarihe dönüşmez): VKN '0911110602', tarih '2026-04-13', No 'TS020...'
TEXT_COLS = {"Fatura Tarihi", "Fatura Numarası", "Vergi Kimlik Numarası"}

HEADER_FILL = "FF1F4E79"


def kaydet_excel(df, save_path, mod):
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    cols = list(TEKLI_COLS if mod == "tekli" else AYRINTILI_COLS)
    widths = TEKLI_WIDTHS if mod == "tekli" else AYRINTILI_WIDTHS

    has_anomali = "Anomali" in df.columns and df["Anomali"].astype(bool).any()
    if has_anomali:
        cols.append("Anomali")

    if "_sheet" not in df.columns:
        df = df.copy()
        df["_sheet"] = "Satış Faturaları" if mod == "tekli" else "Fatura Detayı"

    sheet_order = list(dict.fromkeys(df["_sheet"].tolist()))
    sheet_col = df["_sheet"].values
    df_data = df.drop(columns=["_sheet"])
    df_data = df_data[[c for c in cols if c in df_data.columns]]

    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    anomali_fill = PatternFill("solid", fgColor="FFF2CC")

    def _format_sheet(ws, df_sheet):
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32

        money_idx = {i + 1 for i, c in enumerate(df_sheet.columns) if c in MONEY_COLS}
        text_idx = {i + 1 for i, c in enumerate(df_sheet.columns) if c in TEXT_COLS}
        anomali_idx = (list(df_sheet.columns).index("Anomali") + 1
                       if "Anomali" in df_sheet.columns else None)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            is_anom = bool(ws.cell(row[0].row, anomali_idx).value) if anomali_idx else False
            for cell in row:
                if is_anom:
                    cell.fill = anomali_fill
                if cell.column in money_idx:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif cell.column in text_idx:
                    cell.number_format = "@"          # metin (sıfır/tarih korunur)
                    cell.alignment = Alignment(vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        for i, col_name in enumerate(df_sheet.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 16)
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            dfx = df_data[sheet_col == sheet_name].reset_index(drop=True)
            # Metin sütunlarını gerçekten metin yap (openpyxl'e str ver)
            for tc in TEXT_COLS:
                if tc in dfx.columns:
                    dfx[tc] = dfx[tc].apply(lambda v: "" if v is None else str(v))
            dfx.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            _format_sheet(writer.sheets[sheet_name[:31]], dfx)

        if has_anomali:
            adf = df_data[df_data["Anomali"].astype(bool)].reset_index(drop=True)
            if not adf.empty:
                nm = "⚠ Anomaliler"
                for tc in TEXT_COLS:
                    if tc in adf.columns:
                        adf[tc] = adf[tc].apply(lambda v: "" if v is None else str(v))
                adf.to_excel(writer, index=False, sheet_name=nm)
                _format_sheet(writer.sheets[nm], adf)
                sheet_order.append(nm)

    return len(df_data), len(sheet_order), sheet_order


# ═══════════════════════════════════════════════════════════════════
#  15) GUI
# ═══════════════════════════════════════════════════════════════════

class App(tk.Tk if _GUI_AVAILABLE else object):
    def __init__(self):
        super().__init__()
        self.title("XML Fatura Aktarıcı — Profesyonel")
        self.resizable(True, True)
        self.configure(bg="#1e1e2e")
        self._center()
        self._build_ui()

    def _center(self):
        self.update_idletasks()
        w = 760
        sh = self.winfo_screenheight()
        h = min(900, sh - 60)
        self.geometry(f"{w}x{h}+{(self.winfo_screenwidth() - w) // 2}+{(sh - h) // 2}")

    def _build_ui(self):
        BG, CARD, ACC, ACC2 = "#1e1e2e", "#2a2a3e", "#7c3aed", "#5b21b6"
        FG, MUTED, GREEN, AMBER = "#e2e8f0", "#94a3b8", "#22c55e", "#f59e0b"
        self._c = dict(BG=BG, CARD=CARD, ACC=ACC, ACC2=ACC2, FG=FG,
                       MUTED=MUTED, GREEN=GREEN, AMBER=AMBER)

        hdr = tk.Frame(self, bg=ACC, pady=14)
        hdr.pack(side="top", fill="x")
        tk.Label(hdr, text="📄  XML Fatura Aktarıcı", font=("Segoe UI", 15, "bold"),
                 bg=ACC, fg="white").pack()
        tk.Label(hdr, text="GİB e-Fatura / e-Arşiv XML → Excel  (UBL-TR 1.2)",
                 font=("Segoe UI", 9), bg=ACC, fg="#c4b5fd").pack()

        tk.Frame(self, bg="#3a3a50", height=2).pack(side="bottom", fill="x")
        bot = tk.Frame(self, bg=BG, padx=16, pady=10)
        bot.pack(side="bottom", fill="x")
        self.run_btn = tk.Button(bot, text="▶   AKTARIMI BAŞLAT",
                                 font=("Segoe UI", 12, "bold"), bg=GREEN, fg="white",
                                 relief="flat", pady=12, cursor="hand2", command=self._baslat)
        self.run_btn.pack(fill="x")

        mid = tk.Frame(self, bg=BG)
        mid.pack(side="top", fill="both", expand=True)
        canvas = tk.Canvas(mid, bg=BG, highlightthickness=0, bd=0)
        vsb = tk.Scrollbar(mid, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._scroll_frame = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=self._scroll_frame, anchor="nw")
        self._scroll_frame.bind("<Configure>",
                                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        # ── Çıktı Modu ──
        self._kart("Çıktı Modu", "Faturayı nasıl aktarmak istersiniz?")
        self.mod_var = tk.StringVar(value="tekli")
        bf = tk.Frame(self._ck, bg=CARD)
        bf.pack(fill="x")
        self._rb_tekli = self._radio_btn(
            bf, "Tekli Mod", "Her fatura = 1 satır\nSistemin kabul ettiği 9 sütunluk format",
            "tekli", ACC, CARD)
        self._rb_tekli.pack(side="left", fill="both", expand=True, padx=(0, 6))
        self._rb_det = self._radio_btn(
            bf, "Ayrıntılı Mod", "Her kalem = 1 satır\nBirim fiyat + kalem matrah/KDV",
            "ayrintili", CARD, CARD)
        self._rb_det.pack(side="left", fill="both", expand=True)

        # ── Cins Gruplama ──
        self._card_grp = tk.Frame(self._scroll_frame, bg=CARD, padx=20, pady=14)
        self._card_grp.pack(fill="x", padx=16, pady=(10, 0))
        tk.Label(self._card_grp, text="Cins Gruplama  (Tekli Mod)",
                 font=("Segoe UI", 10, "bold"), bg=CARD, fg=FG).pack(anchor="w")
        tk.Label(self._card_grp, text="Aynı ürünün farklı ebatları nasıl gösterilsin?",
                 font=("Segoe UI", 8), bg=CARD, fg=MUTED).pack(anchor="w", pady=(0, 6))
        self.grp_var = tk.StringVar(value="tam")
        for val, lbl, desc in [
            ("tam", "Tam Ad", "Kalem adı olduğu gibi (önerilen — örnek listeyle aynı)"),
            ("ilk_kelime", "İlk Kelime", "Boru 5×5 → Boru  (hızlı, AI'sız)"),
            ("ai", "Akıllı (AI)", "Claude ile otomatik gruplama"),
        ]:
            f = tk.Frame(self._card_grp, bg=CARD)
            f.pack(anchor="w", pady=2)
            tk.Radiobutton(f, text=lbl, variable=self.grp_var, value=val, bg=CARD, fg=FG,
                           selectcolor=ACC2, activebackground=CARD,
                           font=("Segoe UI", 9, "bold"),
                           command=self._toggle_api_frame).pack(side="left")
            tk.Label(f, text=f"— {desc}", font=("Segoe UI", 8),
                     bg=CARD, fg=MUTED).pack(side="left", padx=4)

        # ── AI Özellikler ──
        self._kart("🤖  AI Destekli Özellikler  (opsiyonel — varsayılan kapalı)",
                   "Seçtiğiniz özellikler aktarım sonrasında çalışır (Anthropic API Key gerekir)")
        self.ai_firma_var = tk.BooleanVar(value=False)
        self.ai_birim_var = tk.BooleanVar(value=False)
        self.ai_xml_var = tk.BooleanVar(value=False)
        self.ai_anomali_var = tk.BooleanVar(value=False)
        for var, baslik, acik in [
            (self.ai_firma_var, "Firma Normalizasyonu",
             "Aynı firmanın farklı yazımlarını birleştirir"),
            (self.ai_birim_var, "Birim Normalizasyonu",
             "Bilinmeyen birim kodlarını çözer  (örn. 'XYZ' → 'ADET')"),
            (self.ai_xml_var, "Bozuk XML Kurtarma",
             "Parse edilemeyen faturaları AI ile kurtarır"),
            (self.ai_anomali_var, "Anomali Tespiti",
             "Şüpheli faturaları işaretler, ⚠ sekmesi ekler"),
        ]:
            f = tk.Frame(self._ck, bg=CARD, pady=4)
            f.pack(fill="x")
            tk.Checkbutton(f, variable=var, bg=CARD, fg=FG, selectcolor=ACC2,
                           activebackground=CARD, font=("Segoe UI", 9, "bold"),
                           command=self._toggle_api_frame, text=baslik).pack(side="left", anchor="n")
            tk.Label(f, text=acik, font=("Segoe UI", 8), bg=CARD, fg=MUTED,
                     justify="left").pack(side="left", padx=(4, 0))

        self._api_frame = tk.Frame(self._ck, bg=CARD, pady=6)
        la = tk.Frame(self._api_frame, bg=CARD)
        la.pack(fill="x", pady=(0, 4))
        tk.Label(la, text="🔑 Anthropic API Key", font=("Segoe UI", 9, "bold"),
                 bg=CARD, fg=AMBER).pack(side="left")
        tk.Label(la, text="  console.anthropic.com → API Keys", font=("Segoe UI", 8),
                 bg=CARD, fg=MUTED).pack(side="left")
        self.api_key_var = tk.StringVar()
        tk.Entry(self._api_frame, textvariable=self.api_key_var, show="*",
                 font=("Consolas", 9), bg="#0f0f1a", fg=FG, insertbackground=FG,
                 relief="flat", width=52).pack(fill="x")

        # ── Dosya Konumları ──
        self._kart("Dosya Konumları", "")
        for lbl, attr, cmd in [("📁 Klasör:", "klasor_var", "_sec_klasor"),
                               ("💾 Kayıt Yeri:", "kayit_var", "_sec_kayit")]:
            row = tk.Frame(self._ck, bg=CARD)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=lbl, font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                     width=13, anchor="w").pack(side="left")
            setattr(self, attr, tk.StringVar(value="Seçilmedi"))
            tk.Label(row, textvariable=getattr(self, attr), font=("Segoe UI", 9),
                     bg=CARD, fg=FG, anchor="w").pack(side="left", fill="x", expand=True)
            tk.Button(row, text="Seç", font=("Segoe UI", 8, "bold"), bg=ACC, fg="white",
                      relief="flat", padx=12, cursor="hand2",
                      command=getattr(self, cmd)).pack(side="right")

        # ── Durum & Log ──
        self._kart("İşlem Durumu", "")
        ph = tk.Frame(self._ck, bg=CARD)
        ph.pack(fill="x")
        self.prog_lbl = tk.Label(ph, text="", font=("Segoe UI", 9), bg=CARD, fg=MUTED)
        self.prog_lbl.pack(side="right")
        self.prog_bar = ttk.Progressbar(self._ck, mode="determinate")
        self.prog_bar.pack(fill="x", pady=(6, 0))
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("TProgressbar", troughcolor=BG, background=ACC,
                    lightcolor=ACC, darkcolor=ACC2, bordercolor=CARD)

        self._kart("Log", "")
        tk.Button(self._ck, text="Temizle", font=("Segoe UI", 8), bg=BG, fg=MUTED,
                  relief="flat", cursor="hand2", command=self._temizle_log).pack(anchor="e")
        self.log_box = scrolledtext.ScrolledText(self._ck, height=7, font=("Consolas", 8),
                                                 bg="#0f0f1a", fg=FG, insertbackground=FG,
                                                 relief="flat", bd=0, state="disabled")
        self.log_box.pack(fill="both", expand=True, pady=(4, 0))
        tk.Frame(self._scroll_frame, bg=BG, height=10).pack()

        self._klasor = ""
        self._kayit = ""
        self._running = False
        self._toggle_api_frame()

    def _kart(self, baslik, alt):
        c = self._c
        f = tk.Frame(self._scroll_frame, bg=c["CARD"], padx=20, pady=14)
        f.pack(fill="x", padx=16, pady=(10, 0))
        if baslik:
            tk.Label(f, text=baslik, font=("Segoe UI", 10, "bold"),
                     bg=c["CARD"], fg=c["FG"]).pack(anchor="w")
        if alt:
            tk.Label(f, text=alt, font=("Segoe UI", 8),
                     bg=c["CARD"], fg=c["MUTED"]).pack(anchor="w", pady=(0, 8))
        self._ck = f

    def _radio_btn(self, parent, title, desc, value, bg_sel, bg_norm):
        c = self._c
        is_sel = self.mod_var.get() == value
        frame = tk.Frame(parent, bg=bg_sel if is_sel else bg_norm, padx=14, pady=10, cursor="hand2")
        tk.Radiobutton(frame, variable=self.mod_var, value=value, bg=frame.cget("bg"),
                       activebackground=frame.cget("bg"),
                       command=self._toggle_mod).pack(side="left", anchor="n")
        inner = tk.Frame(frame, bg=frame.cget("bg"))
        inner.pack(side="left", fill="both", expand=True)
        tk.Label(inner, text=title, font=("Segoe UI", 9, "bold"),
                 bg=frame.cget("bg"), fg=c["FG"], anchor="w").pack(fill="x")
        tk.Label(inner, text=desc, font=("Segoe UI", 8), bg=frame.cget("bg"),
                 fg=c["MUTED"], anchor="w", justify="left").pack(fill="x")
        for w in [frame, inner] + list(inner.winfo_children()):
            w.bind("<Button-1>", lambda e, v=value: [self.mod_var.set(v), self._toggle_mod()])
        return frame

    def _toggle_mod(self):
        c = self._c
        tekli = self.mod_var.get() == "tekli"
        self._rb_tekli.configure(bg=c["ACC"] if tekli else c["CARD"])
        self._rb_det.configure(bg=c["CARD"] if tekli else c["ACC"])
        for w in self._rb_tekli.winfo_children():
            w.configure(bg=c["ACC"] if tekli else c["CARD"])
        for w in self._rb_det.winfo_children():
            w.configure(bg=c["CARD"] if tekli else c["ACC"])
        if tekli:
            self._card_grp.pack(fill="x", padx=16, pady=(10, 0))
        else:
            self._card_grp.pack_forget()
        self._scroll_frame.update_idletasks()

    def _toggle_api_frame(self):
        gerek = (self.grp_var.get() == "ai" or self.ai_firma_var.get() or
                 self.ai_birim_var.get() or self.ai_xml_var.get() or self.ai_anomali_var.get())
        if gerek:
            self._api_frame.pack(fill="x", pady=(8, 0))
        else:
            self._api_frame.pack_forget()

    def _sec_klasor(self):
        p = filedialog.askdirectory(title="Faturaların Olduğu Klasörü Seç")
        if p:
            self._klasor = p
            self.klasor_var.set(p if len(p) < 64 else "..." + p[-61:])

    def _sec_kayit(self):
        default = "Fatura_Tekli.xlsx" if self.mod_var.get() == "tekli" else "Fatura_Ayrintili.xlsx"
        p = filedialog.asksaveasfilename(title="Excel Kayıt Yeri", defaultextension=".xlsx",
                                         initialfile=default, filetypes=[("Excel Dosyası", "*.xlsx")])
        if p:
            self._kayit = p
            self.kayit_var.set(p if len(p) < 64 else "..." + p[-61:])

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _temizle_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _set_progress(self, current, total):
        pct = int(current / total * 100) if total else 0
        self.prog_bar["value"] = pct
        self.prog_lbl.configure(text=f"{current} / {total}  ({pct}%)")
        self.update_idletasks()

    def _baslat(self):
        if self._running:
            return
        if not self._klasor:
            messagebox.showwarning("Uyarı", "Lütfen önce bir klasör seçin.")
            return
        if not self._kayit:
            messagebox.showwarning("Uyarı", "Lütfen kayıt yerini belirtin.")
            return
        gerek = (self.grp_var.get() == "ai" or self.ai_firma_var.get() or
                 self.ai_birim_var.get() or self.ai_xml_var.get() or self.ai_anomali_var.get())
        if gerek and not self.api_key_var.get().strip():
            messagebox.showwarning("API Key Gerekli",
                                   "Seçili AI özellikler için Anthropic API Key girmelisiniz.")
            return
        self._running = True
        self.run_btn.configure(state="disabled", text="⏳  İşleniyor...")
        self._temizle_log()
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        mod = self.mod_var.get()
        grp = self.grp_var.get() if mod == "tekli" else "tam"
        api_key = self.api_key_var.get().strip()
        ai_features = {"gruplama_modu": grp, "firma": self.ai_firma_var.get(),
                       "birim": self.ai_birim_var.get(), "anomali": self.ai_anomali_var.get()}
        ai_xml = self.ai_xml_var.get()

        self._log(f"Mod              : {'Tekli' if mod == 'tekli' else 'Ayrıntılı'}")
        if mod == "tekli":
            etk = {"tam": "Tam Ad", "ilk_kelime": "İlk Kelime", "ai": "Akıllı (AI)"}
            self._log(f"Cins Gruplama    : {etk.get(grp, grp)}")
        self._log(f"Klasör           : {self._klasor}")

        jobs = collect_xml_files(self._klasor)
        self._log(f"Bulunan dosya    : {len(jobs)}\n{'─' * 56}")
        if not jobs:
            self._log("❌ Hiç XML/ZIP dosyası bulunamadı.")
            self._done(False, 0, 0, 0, 0, 0, [])
            return

        if mod == "tekli":
            parse_fn = lambda c, f: parse_tekli(c, f, gruplama_modu=grp, api_key=api_key)
        else:
            parse_fn = parse_ayrintili

        all_data, ok, fail, rec = process_files(
            jobs, parse_fn, self._log, self._set_progress,
            ai_xml_recovery=ai_xml, api_key=api_key, mod=mod)

        self._log(f"\n{'─' * 56}")
        self._log(f"✅ Başarılı: {ok}   🤖 Kurtarılan: {rec}   ❌ Başarısız: {fail}")
        if not all_data:
            self._log("❌ Aktarılacak veri bulunamadı.")
            self._done(False, ok, fail, rec, 0, 0, [])
            return

        anomali = 0
        if (grp == "ai" or ai_features["firma"] or ai_features["birim"]
                or ai_features["anomali"]) and api_key:
            self._log(f"\n{'─' * 56}\n🤖 AI sonrası işlemler başlıyor...")
            all_data, anomali = post_process_ai(all_data, ai_features, api_key, mod, self._log)
        else:
            for row in all_data:
                for k in ("_kalemler", "_ham_birim_kodu", "_mal_cinsi_ham"):
                    row.pop(k, None)

        try:
            df = pd.DataFrame(all_data)
            satir, sekme, liste = kaydet_excel(df, self._kayit, mod)
            self._log(f"\n💾 Excel kaydedildi → {self._kayit}")
            self._log(f"   {satir} satır  |  {sekme} sekme")
            self._done(True, ok, fail, rec, satir, anomali, liste)
        except Exception as e:
            import traceback
            self._log(f"❌ Excel kaydedilemedi: {e}")
            for ln in traceback.format_exc().splitlines():
                self._log(ln)
            self._done(False, ok, fail, rec, 0, anomali, [])

    def _done(self, success, ok, fail, rec, satir, anomali, liste):
        self._running = False
        self.run_btn.configure(state="normal", text="▶   AKTARIMI BAŞLAT")
        if success:
            sekmeler = "\n".join(f"  {'⚠' if '⚠' in s else '📋'} {s}" for s in liste)
            anom = f"\n⚠  Anomali sayısı    : {anomali}" if anomali else ""
            messagebox.showinfo("Tamamlandı ✓",
                                f"İşlem başarıyla tamamlandı!\n\n"
                                f"✅ Başarılı fatura   : {ok}\n"
                                f"🤖 AI kurtarılan    : {rec}\n"
                                f"❌ Başarısız         : {fail}{anom}\n\n"
                                f"📊 Excel satır sayısı: {satir}\n"
                                f"📑 Sekmeler:\n{sekmeler}\n\n"
                                f"Kayıt: {os.path.basename(self._kayit)}")
        else:
            messagebox.showerror("Hata", "İşlem tamamlanamadı.\nLog ekranını kontrol edin.")


# ═══════════════════════════════════════════════════════════════════
#  GİRİŞ NOKTASI
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if not _GUI_AVAILABLE:
        raise SystemExit("Bu program grafik arayüz (tkinter) gerektirir. "
                         "Lütfen masaüstü bir Python kurulumuyla çalıştırın.")
    App().mainloop()
